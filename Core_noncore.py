import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font # type: ignore
import re

def unhide_columns(excel_path, sheet_names):
    wb = load_workbook(excel_path)
    for sheet_name in sheet_names:
        ws = wb[sheet_name]
        for col in ws.column_dimensions:
            ws.column_dimensions[col].hidden = False
    wb.save(excel_path)

def extract_percentage_and_threshold(sob_data, pattern):
    for i, row in sob_data.iterrows():
        text = str(row.iloc[1]).strip()
        match = re.search(pattern, text)
        if match:
            coinsurance_percentage = int(match.group(1).replace(",", ""))  # Extract the percentage
            threshold_value = float(match.group(2).replace(",", "").replace("$", ""))  # Extract the threshold value
            return coinsurance_percentage, threshold_value
    return None, None

def extract_coinsurance_data(sob_data):
    coinsurance_pattern = r'(\d+)% on the 1st \$(\d+(?:,\d{3})*(?:\.\d{2})?)\D+(\d+)% thereafter'
    coinsurance_percentage, threshold_value = extract_percentage_and_threshold(sob_data, coinsurance_pattern)
    if coinsurance_percentage is None or threshold_value is None:
        raise ValueError("Could not find coinsurance data in the SOB sheet.")
    
    smaller_percentage = coinsurance_percentage / 100
    larger_percentage = 1
    stoploss_value = (larger_percentage - smaller_percentage) * threshold_value
    
    return threshold_value, coinsurance_percentage, stoploss_value

def compare_sheet_with_sob(sheet_data, sob_data, sheet_name, user_input):
    benefit_types = ['ANES', 'ASUR', 'CHRT', 'DXL ', 'HAEM', 'HS  ', 'MISC', 'RX  ', 'SURG', 'GAMB','EV  ','OV  ','HOMV','HOSV','SV  ','TOV ','ICU ']
    benefit_types1 = ['ANES', 'ASUR', 'CHRT', 'DXL ', 'HAEM', 'HS  ', 'MISC', 'RX  ', 'SURG', 'GAMB','ICU ']
    coinsurance_percentage_type_3_4 = 40
    
    major_max_text = "For Active Employees under age 65" if user_input == 'yes' else "For Active Employees age 65 & over and Retirees"
    sob_major_max_row = sob_data[sob_data.iloc[:, 0].str.contains(major_max_text, na=False, case=False)]
    if sob_major_max_row.empty:
        raise ValueError(f'Could not find "{major_max_text}" in the SOB sheet.')
    sob_major_max_value = pd.to_numeric(sob_major_max_row.iloc[0, -1], errors='coerce')

    sob_deductible_text = "Per Each Individual Insured"
    sob_deductible_row = sob_data[sob_data.iloc[:, 0].str.contains(sob_deductible_text, na=False, case=False)]
    if sob_deductible_row.empty:
        raise ValueError(f'Could not find "{sob_deductible_text}" in the SOB sheet.')
    sob_deductible_value = pd.to_numeric(sob_deductible_row.iloc[0, -1], errors='coerce')

    sob_familydeductible_text = "Per Family"
    sob_familydeductible_row = sob_data[sob_data.iloc[:, 0].str.contains(sob_familydeductible_text, na=False, case=False)]
    if sob_familydeductible_row.empty:
        raise ValueError(f'Could not find "{sob_familydeductible_text}" in the SOB sheet.')
    sob_familydeductible_value = pd.to_numeric(sob_familydeductible_row.iloc[0, -1], errors='coerce') * sob_deductible_value

    if sheet_name == 'Core':
        threshold, coinsurance_percentage, stoploss_value = extract_coinsurance_data(sob_data)
    else:  # NonCore
        threshold, coinsurance_percentage, stoploss_value = 50000, 75, 50000

    if 'Comments' not in sheet_data.columns:
        sheet_data['Comments'] = ''

    for index, row in sheet_data.iterrows():
        comments = []
        
        if row['Sub-Category'] == 'S':
            core_deductible = pd.to_numeric(row['Individual Deductible'], errors='coerce')
            core_family_deductible = pd.to_numeric(row['Family Deductible'], errors='coerce')
            core_medicalCOin = pd.to_numeric(row['Major Medical  % 1'], errors='coerce')
            core_major_max = pd.to_numeric(row['Major/Base Dollar Max'], errors='coerce')
            core_stoploss = pd.to_numeric(row['STOPLOSS'], errors='coerce')

            if row['Type'] in [1, 2]:
                if row['Benefit Type'] in benefit_types:
                    if core_deductible != sob_deductible_value:
                        comments.append("DEDUCTIBLES SHOULD BE CHANGED")

                    if core_family_deductible != sob_familydeductible_value:
                        if not comments:
                            comments.append(f"FAMILY DEDUCTIBLE SHOULD BE {sob_familydeductible_value} INSTEAD OF {core_family_deductible}")

                    if core_medicalCOin != coinsurance_percentage:
                        comments.append(f"COINSURANCE PERCENTAGE SHOULD BE {coinsurance_percentage}% INSTEAD OF {core_medicalCOin}%")

                    if core_stoploss != round(stoploss_value):
                        comments.append(f"STOPLOSS SHOULD BE {round(stoploss_value)} INSTEAD OF {core_stoploss}")

                    if core_major_max != sob_major_max_value:
                        comments.append(f"LTM SHOULD BE {sob_major_max_value} INSTEAD OF {core_major_max}")

            elif row['Type'] in [3, 4]:
                if row['Benefit Type'] in benefit_types1:
                    if core_deductible != sob_deductible_value:
                        comments.append("DEDUCTIBLES SHOULD BE CHANGED")

                    if core_family_deductible != sob_familydeductible_value:
                        if not comments:
                            comments.append(f"FAMILY DEDUCTIBLE SHOULD BE {sob_familydeductible_value} INSTEAD OF {core_family_deductible}")

                    if core_medicalCOin != coinsurance_percentage_type_3_4:
                        comments.append(f"COINSURANCE PERCENTAGE SHOULD BE {coinsurance_percentage_type_3_4}% INSTEAD OF {core_medicalCOin}%")

                    if core_major_max != sob_major_max_value:
                        comments.append(f"LTM SHOULD BE {sob_major_max_value} INSTEAD OF {core_major_max}")

            if comments:
                sheet_data.at[index, 'Comments'] = ', '.join(comments)

    return sheet_data




def read_comments_and_codes(sheet):
    comments_dict = {}
    valid_codes_dict = {}
    special_benefit_types = ['ANES', 'ASUR', 'CHRT', 'DXL ', 'EV  ', 'GAMB', 'HAEM', 'HOMV', 'HOSV', 
                             'HS  ', 'ICU ', 'MISC', 'OV  ', 'RBO ', 'RX  ', 'SURG', 'SV  ', 
                             'TOV ', 'IMRT', 'SURC', 'ICU ']

    for row in sheet.iter_rows(min_row=2, values_only=True):
        benefit_type, calculation_code, subcategory, relation, comment, type_ = row[2], row[10], row[5], row[4], row[-1], row[6]
        
        if comment:
            key = (benefit_type, calculation_code, subcategory, relation, type_)
            comments_dict[key] = comment.strip() if isinstance(comment, str) else comment

        if subcategory in ['1', '2', '3', 'D', 'O', 'P'] or (subcategory == 'S' and benefit_type not in special_benefit_types) or (type_ == 2):
            code_key = (benefit_type, subcategory, relation)
            valid_codes_dict.setdefault(code_key, set()).add(calculation_code)

    return comments_dict, valid_codes_dict

def update_comments(sheet, comments_dict, valid_codes_dict, local_sheet, sob_data):
    benefits_type_4_match = ['EV  ', 'HOMV', 'HOSV', 'OV  ', 'SV  ', 'TOV ']
    special_benefit_types = ['ANES', 'ASUR', 'CHRT', 'DXL ', 'EV  ', 'GAMB', 'HAEM', 'HOMV', 'HOSV',
                             'HS  ', 'ICU ', 'MISC', 'OV  ', 'RBO ', 'RX  ', 'SURG', 'SV  ',
                             'TOV ', 'IMRT', 'SURC', 'ICU ']
    sob_internal_limit_row = sob_data[sob_data.iloc[:, 0].astype(str).str.contains("Overseas \(Non-Caribbean|Non-Caricom\)", case=False, na=False)]
    sob_internal_limit = pd.to_numeric(sob_internal_limit_row.iloc[0, -1], errors='coerce')

    for row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
        benefit_type = row[2].value
        calculation_code = row[10].value
        subcategory = row[5].value
        relation = row[4].value
        type_ = row[6].value
        internal_limit = row[20].value  # assuming column 21 is internal limit

        if not row[-1].value:  # Only update if comment cell is empty
            local_key = (benefit_type, calculation_code, subcategory, relation, type_)

            #  Case 1: Type 4 with related Type 2 logic
            if type_ == 4 and benefit_type in benefits_type_4_match:
                type2_key = (benefit_type, calculation_code, subcategory, relation, 2)
                if type2_key in comments_dict:
                    row[-1].value = comments_dict[type2_key]
                else:
                    matched = False
                    for local_row in local_sheet.iter_rows(min_row=2, max_col=local_sheet.max_column):
                        if (local_row[2].value == benefit_type and
                            local_row[6].value == 2 and
                            local_row[10].value == calculation_code):
                            matched = True
                            break
                    if not matched:
                        row[-1].value = "CAL CODE SHOULD BE CHANGED"

            #  Case 2: Exact key match in Local comment dict
            elif local_key in comments_dict:
                row[-1].value = comments_dict[local_key]

            #  Case 3: RBO logic
            elif benefit_type == 'RBO ':
                if type_ == 1:
                    m, n, o, t, u, r = row[12].value, row[13].value, row[14].value, row[19].value, row[20].value, row[17].value
                    anes_row = None
                    for s_row in sheet.iter_rows(min_row=2, max_col=sheet.max_column):
                        if s_row[2].value == 'ANES' and s_row[6].value == 1:
                            anes_row = s_row
                            break
                    if anes_row:
                        am, an, ao, at, au = anes_row[12].value, anes_row[13].value, anes_row[14].value, anes_row[19].value, anes_row[20].value
                        if (m == am and n == an and o == ao and t == at and u == au and r == sob_internal_limit):
                            continue  # All good, no comment
                        else:
                            row[-1].value = "CAL CODE SHOULD BE CHANGED"
                
                elif type_ == 3:
                    matched = False
                    for local_row in local_sheet.iter_rows(min_row=2, max_col=local_sheet.max_column):
                        if (local_row[2].value == 'RBL ' and
                            local_row[10].value == calculation_code):
                            row[-1].value = local_row[-1].value
                            matched = True
                            break
                    if not matched:
                        row[-1].value = "CAL CODE SHOULD BE CHANGED"

            #  Case 4: General Type 4 fallback
            elif type_ == 4 and benefit_type != "NUT ":
                if calculation_code in valid_codes_dict.get((benefit_type, subcategory, relation), set()):
                    matched_comment = comments_dict.get(local_key)
                    row[-1].value = matched_comment if matched_comment else 'CAL CODE SHOULD BE CHANGED'

            #  Case 5: Normal cases where benefit_type is not in special list
            elif benefit_type not in special_benefit_types:
                for local_row in local_sheet.iter_rows(min_row=2, max_col=local_sheet.max_column):
                    if (local_row[2].value == benefit_type and
                        local_row[5].value == subcategory and
                        local_row[4].value == relation and
                        local_row[6].value == type_):
                        local_calc_code = local_row[10].value
                        local_comment = str(local_row[-1].value).strip() if local_row[-1].value else ''
                        if calculation_code == local_calc_code:
                            row[-1].value = local_comment
                        else:
                            row[-1].value = 'INVALID CAL CODE'
                        break


        #print(f"Updated comment for row: {row[-1].value}")



def process_data(file_path, user_input='yes'):
    unhide_columns(file_path, ['Core', 'NonCore'])
    
    xl = pd.ExcelFile(file_path)
    core_data = pd.read_excel(xl, sheet_name='Core')
    non_core_data = pd.read_excel(xl, sheet_name='NonCore')
    sob_data = pd.read_excel(xl, sheet_name='SOB')

    updated_core_data = compare_sheet_with_sob(core_data, sob_data, 'Core', user_input)
    updated_non_core_data = compare_sheet_with_sob(non_core_data, sob_data, 'NonCore', user_input)
    
    with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        updated_core_data.to_excel(writer, sheet_name='Core', index=False)
        updated_non_core_data.to_excel(writer, sheet_name='NonCore', index=False)

    wb = load_workbook(file_path)
    local_sheet = wb['Local']
    core_sheet = wb['Core'] if 'Core' in wb.sheetnames else None
    noncore_sheet = wb['NonCore'] if 'NonCore' in wb.sheetnames else None

    comments_dict, valid_codes_dict = read_comments_and_codes(local_sheet)

    if core_sheet:
        update_comments(core_sheet, comments_dict, valid_codes_dict, local_sheet, sob_data)
    if noncore_sheet:
        update_comments(noncore_sheet, comments_dict, valid_codes_dict, local_sheet, sob_data)

    wb.save(file_path)

if __name__ == "__main__":
    file_path = 'processed/processed_data.xlsx'
    user_input = input("Is this comparison for Active Employees (under 65)? Enter 'yes' or 'no': ").strip().lower()
    try:
        process_data(file_path, user_input)
        print(f"Processed data saved to the same file: {file_path}")
    except Exception as e:
        print(f"An error occurred: {e}")
