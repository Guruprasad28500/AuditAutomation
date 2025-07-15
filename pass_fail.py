from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font # type: ignore

def run_pass_fail(processed_filepath='processed/processed_data.xlsx'):
    # Load the processed Excel file
    wb = load_workbook(processed_filepath)
    local_sheet = wb['Local']
    core_sheet = wb['Core'] if 'Core' in wb.sheetnames else None
    noncore_sheet = wb['NonCore'] if 'NonCore' in wb.sheetnames else None
    
    # Define fonts
    bold_font = Font(bold=True)
    red_font = Font(color="FF0000")
    
    # Define the function to insert status column
    def insert_status_column(sheet):
        # Insert a new column before the existing comments column
        max_col = sheet.max_column
        comments_col_index = None
   
        # Determine the index of the comments column
        for idx, cell in enumerate(sheet[1], start=1):
            if cell.value == 'Comments':
                comments_col_index = idx
                break
   
        if comments_col_index is None:
            raise ValueError('No "Comments" column found in the sheet.')
   
        # Insert the new column
        sheet.insert_cols(comments_col_index, 1)
        # Set the header for the new column and apply bold font
        status_header = sheet.cell(row=1, column=comments_col_index, value='PASS/FAIL')
        status_header.font = bold_font
   
        # Update the "Status" column based on the "Comments" column
        for row in sheet.iter_rows(min_row=2, max_col=max_col + 1):
            comments_cell = row[comments_col_index]
            status_cell = row[comments_col_index - 1]
            if comments_cell.value:
                status_cell.value = 'FAIL'
                # Apply red font to all cells with text in the row, including the comments cell
                for cell in row:
                    if cell.value:
                        cell.font = red_font
            else:
                status_cell.value = 'PASS'
    
    # Insert and update the Status column in each sheet
    insert_status_column(local_sheet)
    if core_sheet:
        insert_status_column(core_sheet)
    if noncore_sheet:
        insert_status_column(noncore_sheet)
    
    # Save the updated workbook
    wb.save(processed_filepath)

# If the script is run directly, use the default filepath
if __name__ == "__main__":
    run_pass_fail()
