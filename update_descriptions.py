import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.styles import Font # type: ignore

def run_update_descriptions(processed_filepath='processed/processed_data.xlsx'):
    # Load the workbook and sheets
    wb = load_workbook(processed_filepath)
    local_sheet = wb['Local']
    core_sheet = wb['Core'] if 'Core' in wb.sheetnames else None
    noncore_sheet = wb['NonCore'] if 'NonCore' in wb.sheetnames else None
    
    # Create dictionaries to store comments and valid calculation codes
    comments_dict = {}
    valid_codes_dict = {}
    
    # List of special benefit types to match
    benefits_type_4_match = ['EV  ', 'HOMV', 'HOSV', 'OV  ', 'SV  ','TOV ']
    
    # Special subcategory benefits
    special_benefit_types = ['ANES', 'ASUR', 'CHRT', 'DXL ', 'EV  ', 'GAMB', 'HAEM', 'HOMV', 'HOSV', 'HS  ', 'ICU ', 'MISC', 'OV  ', 'RBO ', 'RX  ', 'SURG', 'SV  ','TOV ','IMRT','SURC']
    
    def read_comments_and_codes(sheet):
        for row in sheet.iter_rows(min_row=2, values_only=True):
            benefit_type, calculation_code, subcategory, relation, comment, type_ = row[2], row[10], row[5], row[4], row[-1], row[6]  # Adjust based on actual column positions
            if comment:
                key = (benefit_type, calculation_code, subcategory, relation, type_)
                comments_dict[key] = comment
    
            # Track valid calculation codes
            if subcategory in ['1', '2', '3', 'D', 'O', 'P'] or (subcategory == 'S' and benefit_type not in special_benefit_types) or (type_ == 2):
                code_key = (benefit_type, subcategory, relation)
                if code_key not in valid_codes_dict:
                    valid_codes_dict[code_key] = set()
                valid_codes_dict[code_key].add(calculation_code)
    
    # Read comments and valid calculation codes from the Local sheet
    read_comments_and_codes(local_sheet)
    
    # Create or clear Description sheet
    if 'Description' not in wb.sheetnames:
        description_sheet = wb.create_sheet('Description')
    else:
        description_sheet = wb['Description']
    
    # Clear existing content in Description sheet
    for row in description_sheet.iter_rows(min_row=1, max_row=description_sheet.max_row):
        for cell in row:
            cell.value = None
    
    # Initialize row positions
    current_row = 2
    bold_font = Font(bold=True)
    
    # Populate the Description sheet with Local comments
    description_sheet.append(['LOCAL COMMENTS'])
    description_sheet.cell(row=current_row, column=1).font = bold_font
    current_row += 1
    added_comments = set()
    for key, comment in comments_dict.items():
        if key[1] in valid_codes_dict.get((key[0], key[2], key[3]), []):
            comment_text = f"{key[0]} - {comment}"
        else:
            comment_text = f"{key[0]} - {comment}"
        if comment_text not in added_comments:
            description_sheet.append([comment_text])
            added_comments.add(comment_text)  # Add to the set to track it as added
            current_row += 1
    
    # Add a gap before Core comments
    description_sheet.append([''])
    current_row += 1
    
    # Populate the Description sheet with Core comments
    if core_sheet:
        description_sheet.append(['CORE COMMENTS'])
        description_sheet.cell(row=current_row, column=1).font = bold_font
        current_row += 1
        added_comments = set()
        for row in core_sheet.iter_rows(min_row=2, values_only=True):
            benefit_type, calculation_code, subcategory, relation, comment = row[2], row[10], row[5], row[4], row[-1]
            if comment:
                comment_text = f"{benefit_type} - {comment}"
                if comment_text not in added_comments:
                    description_sheet.append([comment_text])
                    added_comments.add(comment_text)  # Add to the set to track it as added
                    current_row += 1
    
        # Add a gap before NonCore comments
        description_sheet.append([''])
        current_row += 1
    
    # Populate the Description sheet with NonCore comments
    if noncore_sheet:
        description_sheet.append(['NONCORE COMMENTS'])
        description_sheet.cell(row=current_row, column=1).font = bold_font
        added_comments = set()
        for row in noncore_sheet.iter_rows(min_row=2, values_only=True):
            benefit_type, calculation_code, subcategory, relation, comment = row[2], row[10], row[5], row[4], row[-1]
            if comment:
                comment_text = f"{benefit_type} - {comment}"
                if comment_text not in added_comments:
                    description_sheet.append([comment_text])
                    added_comments.add(comment_text)  # Add to the set to track it as added
                    current_row += 1
    
    # Save the updated workbook
    wb.save(processed_filepath)

# If the script is run directly, use the default filepath
if __name__ == "__main__":
    run_update_descriptions() # type: ignore
