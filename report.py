import pandas as pd # type: ignore
from openpyxl import load_workbook # type: ignore
from openpyxl.worksheet.datavalidation import DataValidation # type: ignore
from openpyxl.styles import Font # type: ignore
 
# Load the processed Excel file
processed_filepath = 'processed/processed_data.xlsx'
 
wb = load_workbook(processed_filepath)
local_sheet = wb['Local']
core_sheet = wb['Core'] if 'Core' in wb.sheetnames else None
noncore_sheet = wb['NonCore'] if 'NonCore' in wb.sheetnames else None
bold_font = Font(bold=True)
# Create the Report sheet if it doesn't exist
if 'Report' not in wb.sheetnames:
    report_sheet = wb.create_sheet('Report')
    # Add headers to the Report sheet
    headers = ['Group No', 'Plan Code', 'Plan Type', 'Benefit Type', 'Relation', 'Subcategory', 'Type', 'Benefit', 'Error Type', 'Financial Type']
    report_sheet.append(headers)
    for cell in report_sheet[1]:
        cell.font = bold_font
else:
    report_sheet = wb['Report']
 
 
# Special subcategory benefits
special_benefit_types = ['ANES', 'ASUR', 'CHRT', 'DXL ', 'EV  ', 'GAMB', 'HAEM', 'HOMV', 'HOSV', 'HS  ', 'ICU ', 'MISC', 'OV  ', 'RBO ', 'RX  ', 'SURG', 'SV  ','TOV ']
def split_comment(comment):
    # Split the comment by ',' or 'and', then strip extra whitespace
    parts = [part.strip() for part in comment.replace('AND ', ',').split(',')]
    return parts
 
def fill_report_from_sheet(sheet, plan_type):
    for row in sheet.iter_rows(min_row=2, values_only=True):
        group_no, plan_code, benefit, relation, subcategory, type_, comment = row[0], row[1], row[2], row[4], row[5], row[6], row[-1]  # Adjust based on actual column positions
 
        if comment:
            # Split the comment into parts if necessary
            comment_parts = split_comment(comment)
            for part in comment_parts:
                error_type = ''
                financial_type = ''
               
                # Determine the error type based on the comment
                if 'DEDUCTIBLES' in part or 'DED' in part:
                    error_type = 'Deductibles'
                elif 'NO OF SERVICES' in part:
                    error_type = 'No of Services'
                elif 'MM PERCENT' in part or 'MM1%' in part or 'COINSURANCE PERCENTAGE' in part:
                    error_type = 'Coinsurance Percentage'
                elif 'STOPLOSS' in part:
                    error_type = 'Stoploss Maximum'
                elif 'LTM' in part or 'LIFE TIME MAXIMUM' in part:
                    error_type = 'Lifetime Maximum'
                elif 'INTERNAL LIMIT' in part or 'WELNESS LIMIT' in part:
                    error_type = 'Internal limit/Benefit limit/Welness limit'
                elif 'BASE COINSURANCE' in part :
                    error_type ='Base Coinsurance'
                elif 'WAITING PERIOD' in part:
                    error_type ='Waiting Period'
                elif 'BENEFIT PERIOD' in part:
                    error_type = 'Benefit Period'
                elif 'ANNUAL MAX' in part or 'ANNUAL MAXIMUM' in  part:
                    error_type ='Annual Maximum'
                elif 'THROUGH PERIOD' in part :
                    error_type = ' Through Period'
                elif part == 'INVALID CAL CODE':
                    error_type = 'Invalid Calculation Code'

                elif 'BASED ON SOB' in part:
                    error_type ='Invalid Caclualtion Code'
                elif 'BENEFIT PERIOD' in part:
                    errro_type ='Benefit Period'
                elif 'ROLLING PERIOD' in part:
                    error_type ='Rolling Period'

               
                # Determine Benefit Type
                if subcategory in ['1', '2', '3', 'S']:
                    benefit_type = "Major Medical"
                elif subcategory in ['D']:
                    benefit_type = "Dental"
                elif subcategory in ['O']:
                    benefit_type = "Vision"
                else:
                    benefit_type = "Preventative"
 
                # Determine Financial Type based on the rules
                if plan_type in ['Core', 'NonCore'] and (subcategory in ['1', '2', '3', 'D', 'O', 'P'] or (subcategory == 'S' and benefit not in special_benefit_types)):
                    financial_type = 'Not Impacted'
 
                if plan_type in ['Core', 'NonCore'] and part == 'CAL CODE SHOULD BE CHANGED':
                    error_type = 'Invalid Calculation Code'
                    financial_type = ''
 
                report_row = [group_no, plan_code, plan_type, benefit_type, relation, subcategory, type_, benefit, error_type, financial_type]
                report_sheet.append(report_row)
 
 
# Fill the Report sheet with data from Local, Core, and NonCore sheets
fill_report_from_sheet(local_sheet, 'Local')
if core_sheet:
    fill_report_from_sheet(core_sheet, 'Core')
if noncore_sheet:
    fill_report_from_sheet(noncore_sheet, 'NonCore')
 
# Apply data validation only if there are rows in the Report sheet
if report_sheet.max_row > 1:  # more than just the header row
    financial_type_options = ['Not Impacted', 'Overpaid', 'Underpaid']
    dv = DataValidation(type="list", formula1=f'"{",".join(financial_type_options)}"', showDropDown=True)
    report_sheet.add_data_validation(dv)
   
    # Apply the data validation to the range
    for row in range(2, report_sheet.max_row + 1):
        dv.add(f'J{row}')
 
# Save the updated workbook
wb.save(processed_filepath)